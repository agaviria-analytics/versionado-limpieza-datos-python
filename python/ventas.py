import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table,TableStyleInfo


archivo='../data/ventas_dia2.xlsx'
df=pd.read_excel(archivo,sheet_name= 'Sheet1')

#Generar un ID
df['ID']=range(1,len(df)+1)

#Reordenar columnas
df=df[["ID","FechaVenta","Zona","Producto","PrecioVenta","Costo"]]

#Reemplazar valos vacios NaN por un texto ejemplo "Por Validar"
df["Zona"]=df["Zona"].fillna("Por validar")
df["Producto"]=df["Producto"].fillna("Por validar")

duplicados_exactos = df[df.duplicated()]
print("Duplicados exactos:\n", duplicados_exactos)
print("Total duplicados exactos:", len(duplicados_exactos))

df["Producto"]

#duplicados_zona_producto=df[df.duplicated(subset=["Zona","Producto"],keep=False)]
#print(duplicados_zona_producto)

df['Zona']=df['Zona'].str.strip().str.lower()
df['Producto']=df['Producto'].str.strip().str.lower()

df['Margen']=df['PrecioVenta']-df['Costo']

print(df['Zona'].value_counts())
print(df['Producto'].value_counts())


def convertir_a_tabla_excel(ruta_archivo, nombre_tabla):
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    col_final = chr(64 + max_col)  # Calcula la última columna (A, B, C...)

    rango = f"A1:{col_final}{max_row}"  # Define el rango A1:col_final
    tabla = Table(displayName=nombre_tabla, ref=rango)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    wb.save(ruta_archivo)
    print(f"✅ Tabla '{nombre_tabla}' creada en {ruta_archivo}")

df.to_excel('../data/archivo_ventas.xlsx', index=False)
convertir_a_tabla_excel('../data/archivo_ventas.xlsx','Tabla_ventas')

print(df)
#print(df.info())
# print(df.dtypes)
# print(df.shape)

